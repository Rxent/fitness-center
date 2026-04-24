"""Экспорт отчёта в PDF (reportlab) и Excel (openpyxl).

Обе функции принимают словарь отчёта в формате, который собирает services.py
(ключи: title, headers, rows, totals), и возвращают готовый HttpResponse —
браузер сразу сохранит файл.
"""
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# Регистрируем системный шрифт DejaVu, чтобы PDF корректно отображал кириллицу.
# DejaVu входит в стандартные Linux-дистрибутивы, а на Windows/Mac
# fallback-шрифт подберётся Python-ом автоматически — если не найдётся,
# просто используем Helvetica (тогда кириллицу писать не сможем, но ошибки не будет).
_CYRILLIC_FONT = 'Helvetica'
_CYRILLIC_FONT_BOLD = 'Helvetica-Bold'
try:
    pdfmetrics.registerFont(
        TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf')
    )
    pdfmetrics.registerFont(
        TTFont(
            'DejaVu-Bold',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        )
    )
    _CYRILLIC_FONT = 'DejaVu'
    _CYRILLIC_FONT_BOLD = 'DejaVu-Bold'
except Exception:
    # На машинах без DejaVu оставим стандартный шрифт — латиница будет,
    # кириллица — скорее всего нет. Проект сам по себе это стерпит.
    pass


def to_excel(report: dict, filename: str) -> HttpResponse:
    """Собирает .xlsx и возвращает его как ответ для скачивания."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Отчёт'

    # Заголовок отчёта в первой строке.
    ws.append([report['title']])
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=max(len(report['headers']), 1))
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Пустая строка для отступа.
    ws.append([])

    # Строка с названиями колонок — синяя заливка, белый жирный шрифт.
    ws.append(report['headers'])
    header_row = ws.max_row
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    for col in range(1, len(report['headers']) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Данные.
    for row in report['rows']:
        ws.append(row)

    # Автоширина колонок (грубо — по максимальной длине значения).
    for idx, _ in enumerate(report['headers'], start=1):
        max_len = len(str(report['headers'][idx - 1]))
        for row in report['rows']:
            if idx - 1 < len(row):
                max_len = max(max_len, len(str(row[idx - 1])))
        ws.column_dimensions[ws.cell(row=header_row, column=idx).column_letter].width = (
            min(max_len + 2, 40)
        )

    # Итоги — отдельным блоком внизу.
    if report.get('totals'):
        ws.append([])
        ws.append(['Итоги'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for name, value in report['totals'].items():
            ws.append([name, value])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def to_pdf(report: dict, filename: str) -> HttpResponse:
    """Собирает .pdf в альбомной ориентации и возвращает его как ответ."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    # Стили: заголовок, подзаголовок, обычный текст — все с кириллическим шрифтом.
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontName=_CYRILLIC_FONT_BOLD,
        alignment=1,
        fontSize=14,
        spaceAfter=10,
    )
    normal = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontName=_CYRILLIC_FONT,
        fontSize=9,
    )

    story = [Paragraph(report['title'], title_style), Spacer(1, 4)]

    # Основная таблица: заголовок + данные.
    data = [report['headers']] + [[str(c) for c in row] for row in report['rows']]
    if len(data) == 1:
        data.append(['Нет данных за выбранный период'] + [''] * (len(report['headers']) - 1))
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), _CYRILLIC_FONT_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), _CYRILLIC_FONT),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    # Итоги под таблицей.
    if report.get('totals'):
        story.append(Spacer(1, 10))
        story.append(Paragraph('<b>Итоги:</b>', normal))
        for name, value in report['totals'].items():
            story.append(Paragraph(f'{name}: <b>{value}</b>', normal))

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response
